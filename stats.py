#! python3

import openpyxl, time
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

start_time = time.time()

district_names = ['AN', 'AS', 'CTX']

districts = {'AN' :  ['TX-AUSTIN ANDERSON ARBOR',
                      'TX-AUSTIN CEDAR PARK',
                      'TX-AUSTIN CYPRESS CREEK',
                      'TX-AUSTIN HESTERS CROSSING',
                      'TX-AUSTIN NORTH ROUND ROCK',
                      'TX-AUSTIN TECHRIDGE',
                      'TX-GEORGETOWN',
                      'TX-PFLUGERVILLE'],
             'AS' : ['TX-AUSTIN BEE CAVES',
                      'TX-AUSTIN BELTERRA',
                      'TX-AUSTIN DOWNTOWN 6THBRAZOS',
                      'TX-AUSTIN HIGHLAND',
                      'TX-AUSTIN NORTH',
                      'TX-AUSTIN SOUTH',
                      'TX-AUSTIN SOUTH CENTRAL',
                      'TX-AUSTIN SOUTHEAST',
                      'TX-AUSTIN WESTLAKE',
                      'TX-SAN MARCOS'],
             'CTX' : ['TX-BELLMEAD',
                      'TX-BRYAN',
                      'TX-COLLEGE STATION',
                      'TX-COPPERAS COVE',
                      'TX-KILLEEN',
                      'TX-TEMPLE',
                      'TX-TOWER POINT',
                      'TX-VICTORIA',
                      'TX-WACO'],
             'COL' : ['CO-BRIARGATE',
                      'CO-RUSTIC HILLS',
                      'CO-SECURITY'],
             'DAL' : ['TX-DALLAS UPTOWN',
                      'TX-DESOTO',
                      'TX-PRESTON CENTER',
                      'TX-RICHARDSON',
                      'TX-WAXAHACHIE'],
             'OK' : ['OK-DEL CITY',
                     'OK-EDMUND',
                     'OK-MOORE',
                     'OK-NORMAN',
                     'OK-NORTHWEST',
                     'OK-PENN CROSSING',
                     'OK-QUAIL SPRINGS',
                     'OK-WEST EDMUND',
                     'OK-YUKON',
                     'TX-WICHITA FALLS'],
             'CAL' : ['CA-LONG BEACH',
                      'CA-PASADENA',
                      'CA-VENICE'],
             'SAC' : ['TX-SAN ANTONIO BANDERA POINTE',
                      'TX-SAN ANTONIO BANDERA TRAILS',
                      'TX-SAN ANTONIO MEDICAL CENTER',
                      'TX-SAN ANTONIO TEZEL'],
             'SAN' : ['TX-NEW BRAUNFELS',
                      'TX-SAN ANTONIO 78THWALZEM',
                      'TX-SAN ANTONIO ALAMO HEIGHTS',
                      'TX-SAN ANTONIO BULVERDE',
                      'TX-SAN ANTONIO EVANS ROAD',
                      'TX-SAN ANTONIO HILL COUNTRY VILLAGE',
                      'TX-SAN ANTONIO LIVE OAK',
                      'TX-SAN ANTONIO LOCKHILL VILLAGE',
                      'TX-SAN ANTONIO NACO',
                      'TX-SAN ANTONIO THE QUARRY',
                      'TX-SAN ANTONIO TRAVIS'],
             'SAS' : ['TX-BROWNSVILLE',
                      'TX-SAN ANTONIO BABCOCK',
                      'TX-SAN ANTONIO CROSSROADS',
                      'TX-SAN ANTONIO CULEBRA',
                      'TX-SAN ANTONIO FIESTA TRAILS',
                      'TX-SAN ANTONIO GOLIAD',
                      'TX-SAN ANTONIO LEGACY TRAILS',
                      'TX-SAN ANTONIO MILITARY',
                      'TX-SAN ANTONIO PRUE ROAD',
                      'TX-SAN ANTONIO ROGERS RANCH',
                      'TX-SAN ANTONIO VALLEY HI',
                      'TX-SOUTH STAPLES']
             }

def loadGym(GYMS):
    dictionary = {}
    for i in range(len(GYMS)):
        dictionary.update({GYMS[i] : {}})
    return dictionary

def lastFirst(firstlast):
    if firstlast == 'Marziyeh Nasiri Shoja':
        lastFirst = 'Nasiri Shoja, Marziyeh'
    else:
        name = firstlast.split(' ')
        last = name.pop()
        name.insert(0, last)
        name[0] = name[0] + ','
        lastFirst = ' '.join(name)
    return lastFirst

def noDollar(number):
    if number:
        number = str(number)
        number = float(number.replace('$', ''))
    else:
        number = 0
    return number


def districtStats(GYMS):
    #Gold's Gym Experience Report
    print('Calculating Gym Data and Sales Data...')
    GGEWB = openpyxl.load_workbook('Golds Gym Experience Report with Detail.xlsx') 
    ggesheet = GGEWB['New Agreements Detail ']

    GymData = loadGym(GYMS)
    for i in range(len(GymData)):
        GymData[GYMS[i]].update( {'POS NBU' : 0, 'NMC' : 0, 'POS FP Set' : 0, 'POS AA' : 0, 'SS' : 0, 'Clients': 0, 'Frequency': 0, 'FP Set' : 0, 'FP Show' : 0, 'APOS NBU' : 0, 'FP Setup' : 0} )
           
    SalesPeopleData = loadGym(GYMS)

    PTSalesData = loadGym(GYMS)


    gymName = column_index_from_string('A')
    isNM = column_index_from_string('V')
    isDB = column_index_from_string('G')
    isRet = column_index_from_string('R')
    isFAO = column_index_from_string('J')
    FP_y = column_index_from_string('X')
    AA_tier = column_index_from_string('K')
    salesPerson = column_index_from_string('AE')
    noMem = ['Staff', 'Trade', 'Lead']

    for i in range(len(GYMS)):
        for row in ggesheet.rows:
            NM = row[isNM-1].value
            DB = row[isDB-1].value
            Ret = row[isRet-1].value
            FAO = row[isFAO-1].value
            tier = row[AA_tier-1].value
            sales_person = row[salesPerson-1].value
            if row[gymName-1].value == GYMS[i] and NM > 0:                                        
                if DB.startswith('z') or any(x in (DB) for x in noMem):     
                    continue
                else:
                    GymData[GYMS[i]]['NMC'] += 1
                    if sales_person in SalesPeopleData[GYMS[i]]:
                        SalesPeopleData[GYMS[i]][sales_person]['NMC'] += 1
                        if 'FAO' in FAO:
                            SalesPeopleData[GYMS[i]][sales_person]['Mem Unit'] += 0.75 #if it catches an FAO in column J, it's a 0.75 no matter what and should ground out
                        elif 'Retail' in Ret:
                            SalesPeopleData[GYMS[i]][sales_person]['Mem Unit'] += 1 #if it doesn't catch an FAO, it should see a retail in column R
                        else:
                            SalesPeopleData[GYMS[i]][sales_person]['Mem Unit'] += 0.75 #everything that isn't a retail (such as corp), would ground out as a 0.75
                    else:
                        if 'FAO' in FAO:
                            SalesPeopleData[GYMS[i]].update( {sales_person  : {'NMC' : 1, 'FP' : 0, 'AA' : 0, 'PT' : 0, 'Rev' : 0, 'Mem Unit' : 0.75, 'AA Unit' : 0}} )
                        elif 'Retail' in Ret:
                            SalesPeopleData[GYMS[i]].update( {sales_person  : {'NMC' : 1, 'FP' : 0, 'AA' : 0, 'PT' : 0, 'Rev' : 0, 'Mem Unit' : 1, 'AA Unit' : 0}} )        
                        else:
                            SalesPeopleData[GYMS[i]].update( {sales_person  : {'NMC' : 1, 'FP' : 0, 'AA' : 0, 'PT' : 0, 'Rev' : 0, 'Mem Unit' : 0.75, 'AA Unit' : 0}} )

                if row[FP_y-1].value:                                                                                                
                    SalesPeopleData[GYMS[i]][sales_person]['FP'] += 1                                                    
                    GymData[GYMS[i]]['POS FP Set'] += 1
                    
                if tier != 'Access':                                        
                    SalesPeopleData[GYMS[i]][sales_person]['AA'] += 1
                    GymData[GYMS[i]]['POS AA'] += 1
                    if tier != 'GoldsPTx/Mo':
                        if 'Enhanced' in tier:
                            SalesPeopleData[GYMS[i]][sales_person]['AA Unit'] += 0.25
                        elif 'Bootcamp' or 'Studio' in tier:
                            SalesPeopleData[GYMS[i]][sales_person]['AA Unit'] += 0.5
                        else:
                            continue

    #Mem Unit: 0, AA Unit: 0
        #if retail mem unit 1
        #if GR/FAO mem unit 0.75
        #if enhanced aa unit: 0.25
        #if bc/st aa unit: 0.5
        #if 3 pack: aa unit: 1
                        
    #PT Sessions Serviced Report
    ptTrainingWB = openpyxl.load_workbook('PT Business Report - PT Sessions Serviced with Date Range.xlsx')
    gsheet = ptTrainingWB['PT Sessions Serviced Summary']
    tsheet = ptTrainingWB['PT Sessions Serviced Individual']

    gGymName = column_index_from_string('D')
    gClients = column_index_from_string('F')
    gFreq = column_index_from_string('J')
    gSes = column_index_from_string('H')
                                          
    tGymName = column_index_from_string('E')
    ptName = column_index_from_string('F')
    tClients = column_index_from_string('G')
    tFreq = column_index_from_string('K')
    tSes = column_index_from_string('I')

    for i in range(len(GYMS)):
        for row in gsheet.rows:
            clients = row[gClients-1].value
            freq = row[gFreq-1].value
            ses = row[gSes-1].value
            if row[gGymName-1].value == GYMS[i]:
                GymData[GYMS[i]]['SS'] = ses
                GymData[GYMS[i]]['Clients'] = clients
                GymData[GYMS[i]]['Frequency'] = freq

    for i in range(len(GYMS)):
        for row in tsheet.rows:
            clients = row[tClients-1].value
            freq = row[tFreq-1].value
            ses = row[tSes-1].value
            if row[tGymName-1].value == GYMS[i]:
                if row[ptName-1].value in PTSalesData[GYMS[i]]:
                    PTSalesData[GYMS[i]][row[ptName-1].value]['Clients'] = clients
                    PTSalesData[GYMS[i]][row[ptName-1].value]['Sessions'] = ses
                    PTSalesData[GYMS[i]][row[ptName-1].value]['Frequency'] = freq                      
                else:
                    PTSalesData[GYMS[i]].update( {row[ptName-1].value : {'Sessions' : 0, 'Classes' : 0, 'Clients' : 0, 'Frequency' : 0, 'Set' : 0, 'Show' : 0, 'Close' : 0, 'PT New Rev': 0, 'PT Renew Rev': 0}} )        #if they're not listed, add them with some stats
                    PTSalesData[GYMS[i]][row[ptName-1].value]['Clients'] = clients
                    PTSalesData[GYMS[i]][row[ptName-1].value]['Sessions'] = ses
                    PTSalesData[GYMS[i]][row[ptName-1].value]['Frequency'] = freq
    

    print('Calculating PT Sales Data...')
    #PT Sales Report
    ptSalesWB = openpyxl.load_workbook('PT Business Report - PT Sales.xlsx') #open 'pt sales report.xlsx'
    sheet = ptSalesWB['PT Business Report - PT Sales']
    #Service Provider Activity Summary
    ptfpWB = openpyxl.load_workbook('Service Provider Activity Summary.xlsx')
    fpsheet = ptfpWB['Sheet1']

    gymName = column_index_from_string('A')
    provider = column_index_from_string('B')
    show = column_index_from_string('I')
    service = column_index_from_string('H')
    fp_type = ['GOLD\'S 3D', 'Fitness Profile', 'Fitness Profile Follow-Up', 'Fit Profile', 'Fit Profile Follow-Up', 'Fitness Assessment']
    for i in range(len(GYMS)):          #cycle through the number of times for each gym named in GYMS                                                   
        for row in fpsheet.rows:       #go down each of the rows in the sheet
            if row[gymName-1].value == GYMS[i] and row[provider-1].value != 'No Service Provider' and row[provider-1].value:
                if any(x in (row[service-1].value) for x in fp_type):
                    pt = lastFirst(row[provider-1].value)
                    if pt in PTSalesData[GYMS[i]]:                                         #if the salesperson in column 'AE' is listed in SalesPeopleData of that gym
                        PTSalesData[GYMS[i]][pt]['Set'] += 1                                 #give that salesperson a +1 nmc to NMC dictionary
                        GymData[GYMS[i]]['FP Set'] += 1
                    else:
                        PTSalesData[GYMS[i]].update( {pt : {'Sessions' : 0, 'Classes' : 0, 'Clients' : 0, 'Frequency' : 0, 'Set' : 1, 'Show' : 0, 'Close' : 0, 'PT New Rev': 0, 'PT Renew Rev' : 0}} )        #if they're not listed, add them with some stats
                        GymData[GYMS[i]]['FP Set'] += 1
                    if row[show-1].value:                                                                                                #if they set an FP:
                        PTSalesData[GYMS[i]][pt]['Show'] += 1                                                    #record it in SalesPeopleData[their gym][their name][fp]
                        GymData[GYMS[i]]['FP Show'] += 1

    PTGymSales = loadGym(GYMS)

    gymName = column_index_from_string('D')
    nbuColumn = column_index_from_string('AG')
    salesPerson = column_index_from_string('F')
    department_col = column_index_from_string('G')
    package_name = column_index_from_string('L')
    pos_y = column_index_from_string('N')
    paid = column_index_from_string('X')
    last_act = column_index_from_string('AD')
    pt_department = ['Asst Fitness Manager', 'Fitness Advisor', 'Fitness Director', 'Fitness Svc Manager 1', 'Fitness Svc Manager 2', 'Fitness Svc Manager 3', 'PT Level 1', 'PT Level 2', 'PT Level 3', 'PT Level 4', 'Studio Coach']
    sales_department = ['Membership Advisor', 'Asst General Manger', 'DM/SVP', 'Front Desk Associate', 'Front Desk Lead', 'General Manager']

    for i in range(len(GYMS)):                                                          #for the number of clubs in GYMS
        for row in sheet.rows:                                                          #look through each row in the sheet's rows  
            gym = row[gymName-1].value
            sales_person = row[salesPerson-1].value
            amount = row[paid-1].value
            department = row[department_col-1].value
            intro = row[package_name-1].value
            pos = row[pos_y-1].value
            nbu = row[nbuColumn-1].value
            last_pt = row[last_act-1].value

            if gym == GYMS[i] and nbu == 'Y':      
                if sales_person in SalesPeopleData[GYMS[i]] and any(x in (department) for x in sales_department):                   
                    SalesPeopleData[GYMS[i]][sales_person]['PT'] += 1                                                                           
                    SalesPeopleData[GYMS[i]][sales_person]['Rev'] += noDollar(amount)
                    if 'INTRO' in intro:
                        SalesPeopleData[GYMS[i]][sales_person]['AA Unit'] += 1
                    if 'POS' in pos:
                        GymData[GYMS[i]]['POS NBU'] += 1
                    else:
                        GymData[GYMS[i]]['APOS NBU'] += 1
                elif sales_person in PTSalesData[GYMS[i]] and any(x in (department) for x in pt_department):
                    PTSalesData[GYMS[i]][sales_person]['Close'] += 1
                    PTSalesData[GYMS[i]][sales_person]['PT New Rev'] += noDollar(amount)
                    if 'POS' in pos:
                        GymData[GYMS[i]]['POS NBU'] += 1
                    else:
                        GymData[GYMS[i]]['APOS NBU'] += 1
                elif sales_person not in PTSalesData[GYMS[i]] and any(x in (department) for x in pt_department):
                    PTSalesData[GYMS[i]].update( {sales_person : {'Sessions' : 0, 'Classes' : 0, 'Clients' : 0, 'Frequency' : 0, 'Set' : 0, 'Show' : 0, 'Close' : 1, 'PT New Rev' : 0, 'PT Renew Rev' : 0}} )
                    PTSalesData[GYMS[i]][sales_person]['PT New Rev'] += noDollar(amount)
                    if 'POS' in pos:
                        GymData[GYMS[i]]['POS NBU'] += 1
                    else:
                        GymData[GYMS[i]]['APOS NBU'] += 1
                else:
                    continue
                        
            elif gym == GYMS[i] and nbu == 'N' and department:
                if last_pt:
                    if last_pt not in PTSalesData[GYMS[i]]:
                        PTSalesData[GYMS[i]].update( {last_pt : {'Sessions' : 0, 'Classes' : 0, 'Clients' : 0, 'Frequency' : 0, 'Set' : 0, 'Show' : 0, 'Close' : 0, 'PT New Rev' : 0, 'PT Renew Rev' : 0}} )
                        PTSalesData[GYMS[i]][last_pt]['PT Renew Rev'] += noDollar(amount)
                    else:
                        PTSalesData[GYMS[i]][last_pt]['PT Renew Rev'] += noDollar(amount)
                elif sales_person not in PTSalesData[GYMS[i]] and any(x in (department) for x in pt_department) and not last_pt:
                    PTSalesData[GYMS[i]].update( {sales_person : {'Sessions' : 0, 'Classes' : 0, 'Clients' : 0, 'Frequency' : 0, 'Set' : 0, 'Show' : 0, 'Close' : 0, 'PT New Rev': 0, 'PT Renew Rev': 0}} )        #if they're not listed, add them with some stats
                    PTSalesData[GYMS[i]][sales_person]['PT Renew Rev'] += noDollar(amount)
                elif sales_person in PTSalesData[GYMS[i]] and any(x in (department) for x in pt_department) and not last_pt:  
                    PTSalesData[GYMS[i]][sales_person]['PT Renew Rev'] += noDollar(amount)
                else:
                    continue
                
            else:
                continue
 

    #Daily Service Provider Scheduler
    print('Calculating Classes...')
    classesWB = openpyxl.load_workbook('Daily Service Provider Scheduler.xlsx')
    csheet = classesWB['Sheet1']

    club_name = column_index_from_string('C')
    service_provider = column_index_from_string('A')
    event = column_index_from_string('U')
    attendance = column_index_from_string('V')
    studio = ['BOOTCAMP', 'GOLD\'S FIT', 'GOLD\'S CYCLE', 'GOLD\'S CYCLE BEATS', 'GOLD\'S CYCLE', 'STUDIO FUSION', 'GOLD\'S BURN']

    for i in range(len(GYMS)):
        for row in csheet.rows:
            gym_name = row[club_name-1].value
            instructor = row[service_provider-1].value
            classes = row[event-1].value
            attendees = row[attendance-1].value
            if instructor in PTSalesData[GYMS[i]] and classes:
                #employees[instructor]['gym'] = gym_name
                if any(x in (classes) for x in studio) and attendees > 0:
                    PTSalesData[GYMS[i]][instructor]['Classes'] += 1
                else:
                    PTSalesData[GYMS[i]][instructor]['Classes'] += 1
            else:
                continue
    
    
    #Member Appointments
    print('Calculating FP Setup...')
    setupWB = openpyxl.load_workbook('Member Appointments.xlsx')
    appsheet = setupWB['Sheet1']
    
    gymName = column_index_from_string('B')
    fp_col = column_index_from_string('D')
    fp_type = ['3D Scan', 'GOLD\'S 3D', 'Fitness Profile', 'Fitness Profile Follow-Up', 'Fit Profile', 'Fit Profile Follow-Up', 'Fitness Assessment']

    for i in range(len(GYMS)):
        for row in appsheet.rows:
            fp = row[fp_col-1].value
            if row[gymName-1].value == GYMS[i] and any(x in (fp) for x in fp_type):
                GymData[GYMS[i]]['FP Setup'] += 1 
                

    #write to new file
    print('Writing...')
    testWB = Workbook()
    sheet = testWB.active
    testWB.remove(sheet)

    #Write district's gym data
    testWB.create_sheet('District')
    testSheet = testWB['District']

    #           A-1       B-2    C-3      D-4          E-5      F-6       G-7       H-8          I-9      J-10       K-11         L-12        M-13        N-14         O-15    P-16     Q-17
    headers = ['Gym', 'POS NBU', 'NMC', 'POS FP Set', 'POS AA', 'SS',  'Clients', 'Frequency', 'FP Set', 'FP Show', 'APOS NBU', 'FP Setup', 'FP Show %', 'FP Close %', 'FP %', 'AA %', 'FP:NBU']
       
    for i in range(len(headers)):
        testSheet.cell(row=1, column=i+1).font = Font(bold=True)
        testSheet.cell(row=1, column=i+1).value = headers[i]

    row = 2
    for gym, stats in GymData.items():
        testSheet.cell(row=row, column=1, value=gym)
        column = 2
        for stats, numbers in stats.items():
            testSheet.cell(row=row, column=column, value=numbers)
            column += 1
        testSheet.cell(row=row, column=column).value = '=J' + str(row) + '/I' + str(row) # FP show %
        column += 1
        testSheet.cell(row=row, column=column).value = '=K' + str(row) + '/J' + str(row)  # FP close %
        column += 1
        testSheet.cell(row=row, column=column).value = '=I' + str(row) + '/C' + str(row)  # FP set % at POS
        column += 1
        testSheet.cell(row=row, column=column).value = '=E' + str(row) + '/C' + str(row)  # AA % at POS
        column += 1
        testSheet.cell(row=row, column=column).value = '=I' +str(row) + '/K' + str(row)  #FP set need for NBU
        row += 1


    

    #Write Sales Data
    #                 A-1       B-2          C-3    D-4  E-5      F-6       G-7       H-8         I-9      J-10     K-11 
    sales_headers = ['Gym', 'Sales Person', 'NMC', 'FP', 'AA', 'PT NBU', 'Revenue', 'Mem Unit', 'AA Unit', 'FP %', 'AA %']

    testWB.create_sheet('Sales')
    testSheet = testWB['Sales']

    for i in range(len(sales_headers)):
        testSheet.cell(row=1, column=i+1).font = Font(bold=True)
        testSheet.cell(row=1, column=i+1).value = sales_headers[i]
    row = 2
    for gym,team in SalesPeopleData.items():  # nmc, fp, aa, pt nbu, revenue, munit, aunit, fp %, aa %
        for salesperson,stats in team.items():  # for salsperson : stats in team,
            testSheet.cell(row=row, column=1, value=gym)
            testSheet.cell(row=row, column=2, value=salesperson)
            column = 3
            for stat,num in stats.items(): 
                testSheet.cell(row=row, column=column, value=num) 
                column += 1
            testSheet.cell(row=row, column=column).value = '=D' + str(row) + '/C' + str(row)  # FP set % 
            column += 1
            testSheet.cell(row=row, column=column).value = '=E' + str(row) + '/C' + str(row)   # AA %
            row += 1
        for i in range(len(sales_headers)):
            testSheet.cell(row=row, column=i+1).font = Font(bold=True)
            testSheet.cell(row=row, column=i+1).value = sales_headers[i]
        row += 1


    testWB.create_sheet('PT')
    testSheet = testWB['PT']
    #Write PT Data

    #               A-1       B-2        C-3      D-4       E-5           F-6       G-7     H-8     I-9        J-10      K-11      L-12      M-13
    pt_headers = ['Gym', 'PT Name', 'Sessions', 'Classes', 'Clients', 'Frequency', 'Set', 'Show', 'Close', 'New Rev', 'Renew Rev', 'Show %', 'Close %']

    for i in range(len(pt_headers)):
        testSheet.cell(row=1, column=i+1).font = Font(bold=True)
        testSheet.cell(row=1, column=i+1).value = pt_headers[i]
    row = 2
    for gym,team in PTSalesData.items():
        for pt,stats in team.items():
            testSheet.cell(row=row, column=1, value=gym)
            testSheet.cell(row=row, column=2, value=pt)
            column = 3
            for stat,num in stats.items(): #set/show/close/pt new rev/pt renewal/show percent/close percent
                testSheet.cell(row=row, column=column, value=num) 
                column += 1
            testSheet.cell(row=row, column=column).value = '=H' + str(row) + '/G' + str(row) #show %
            column += 1
            testSheet.cell(row=row, column=column).value = '=I' + str(row) + '/H' + str(row) #close %
            row += 1
        for i in range(len(pt_headers)):
            testSheet.cell(row=row, column=i+1).font = Font(bold=True)
            testSheet.cell(row=row, column=i+1).value = pt_headers[i]
        row += 1

        
    print('Saving...')
    file_name = 'gym' + str(file_num) + '.xlsx'
    testWB.save(file_name)
    testWB.close()
    print('Done.')

file_num = 1
for i in range(len(district_names)):
    districtStats(districts[district_names[i]])
    print("--- %s seconds ---" % (time.time() - start_time))
    file_num += 1

